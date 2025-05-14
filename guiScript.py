import csv
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from copy import copy
import datetime
import dearpygui.dearpygui as dpg

#-------------------------------------------------------------------------------

badDrvDrawingNum = False
badMainPcbNum = False
badCpuPcbNum = False
badPowerPcbNum = False
badFilterPcbNum = False

allSingleDrives = []
allDualDrives = []
allPowerSupplies = []

dpg.create_context()

#-------------------------------------------------------------------------------

def serialize(sender, app_data):

    #########################################################################
    #Hide / show error text, depending on if bad num flags are present

    drvInfo = []
    [drvInfo, badDrvDrawingNum, badMainPcbNum,
     badCpuPcbNum, badPowerPcbNum, badFilterPcbNum] = generateDrvSerialNumber(dpg.get_value(drvNumber))

    if badDrvDrawingNum:
        dpg.show_item(bDDN)
    elif dpg.is_item_shown(bDDN):
        dpg.hide_item(bDDN)
    
    if badMainPcbNum:
        dpg.show_item(bMPN)
    elif dpg.is_item_shown(bMPN):
        dpg.hide_item(bMPN)
   
    if badCpuPcbNum:
        dpg.show_item(bCPN)
    elif dpg.is_item_shown(bCPN):
        dpg.hide_item(bCPN)

    if badPowerPcbNum:
        dpg.show_item(bPPN)
    elif dpg.is_item_shown(bPPN):
        dpg.hide_item(bPPN)
    
    if badFilterPcbNum:
        dpg.show_item(bFPN)
    elif dpg.is_item_shown(bFPN):
        dpg.hide_item(bFPN)

    #############################################################
    #If the sequential number is unique, add drive to allSingleDrives and add SN to window

    if not (badDrvDrawingNum or badMainPcbNum or badCpuPcbNum or badPowerPcbNum or badFilterPcbNum):

        bSQV = False        
        for row in allSingleDrives:
            if (int(row[9][-4:]) == int(dpg.get_value(seqNum))):
                bSQV = True
                break

        if not bSQV:
            allSingleDrives.append(drvInfo)
            dpg.add_text(
                default_value=drvInfo[9],
                parent=snWindow,
                tag=dpg.get_value(seqNum) 
            )
            dpg.hide_item(bSQN)
        else:
            dpg.show_item(bSQN)
            
    return

#-------------------------------------------------------------------------------

def getDriveInfo(drvNum):

    ##############################################################
    #load the reference CSV file

    with open(r'Data/driveNumbers.csv', 'r', newline='', encoding='utf-8-sig') as csvfile:
        csv_reader = csv.reader(csvfile, delimiter=',', quotechar='|')
        driveNumbers = list(csv_reader)
        csvfile.close()
        
    ##############################################################
    #Loop through all entries in the file

    drvInfo = ['43687REVC','650VDC','inverted','25A','single']
    f = 0
    
    for row in driveNumbers:
        if row[0] == drvNum:
            drvInfo = row
            f = 1
            break
        
    if f == 1:      #if # is recognized, unset flag
        badDrvDrawingNum = False
    else:           #if # is NOT recognized, setting flag that brings up error text
        badDrvDrawingNum = True

    return [drvInfo, badDrvDrawingNum]

#-------------------------------------------------------------------------------

def generateDrvSerialNumber(drvNum):

##############################################################
#get the drive number

    [drvInfo, badDrvDrawingNum] = getDriveInfo(drvNum)
    
##############################################################
#get the pcb numbers, depending on whether drive is dual or single node

    
    boardNums = [dpg.get_value(mainPCB), dpg.get_value(cpuPCB),
                 dpg.get_value(powerPCB), dpg.get_value(filterPCB)]
    drvInfo.extend(boardNums)
    
    if drvInfo[2] == "inverted":                            #first two places
        sn = "DG"
    elif drvInfo[2] == "upright":
        sn = "DE"

    if drvInfo[1] == "650VDC" and drvInfo[3] == "13A":      #3-4 places
        sn = sn + "10"                                      
    elif drvInfo[1] == "650VDC" and drvInfo[3] == "25A":
        sn = sn + "06"
    elif drvInfo[1] == "650VDC" and drvInfo[3] == "50A":
        sn = sn + "13"
    elif drvInfo[1] == "325VDC" and drvInfo[3] == "25A":
        sn = sn + "14"
    elif drvInfo[1] == "325VDC" and drvInfo[3] == "50A":
        sn = sn + "12"
    elif drvInfo[1] == "325VDC" and drvInfo[3] == "100A":
        sn = sn + "11"

    [snPartA, drvInfo[5], badMainPcbNum] = serializeFromCSV(r'Data\snDriveMainBoard.csv', 'Main', drvInfo[5]) #5-6
    sn = sn + snPartA

    [snPartB, drvInfo[6], badCpuPcbNum] = serializeFromCSV(r'Data\snDriveCPUBoard.csv', 'CPU', drvInfo[6]) #7
    sn = sn + snPartB

    [snPartC, drvInfo[7], badPowerPcbNum] = serializeFromCSV(r'Data\snDrivePowerBoard.csv', 'Power', drvInfo[7]) #8-9
    sn = sn + snPartC
        
    [snPartD, drvInfo[8], badFilterPcbNum] = serializeFromCSV(r'Data\snDriveFilterBoard.csv','Filter', drvInfo[8]) #10
    sn = sn + snPartD

    sn = sn +'-'+ dpg.get_value(seqNum)  
    drvInfo.append(str(sn))
        
    return [drvInfo, badDrvDrawingNum, badMainPcbNum, badCpuPcbNum, badPowerPcbNum, badFilterPcbNum]

#----------------------------------------------------------------------------

def serializeFromCSV(filename, pcbType, pcb2SN):       #string filename, pcb to serialize (####X-####), 
                                                      #type of PCB for print statements
##############################################################
#load the reference CSV file

    with open(filename, 'r', newline='', encoding='utf-8-sig') as csvfile:
        csv_reader = csv.reader(csvfile, delimiter=',', quotechar='|')
        boards = list(csv_reader)
    csvfile.close()
            
##############################################################
#Loop through all entries in the file, if number isn't recognized, 
#loop through asking for number to be re-entered or to add a new entry to the file

    snPart = '0'
    f = 0    

    for row in boards:
        if row[0] == pcb2SN[0:-5]:
            snPart = row[1]
            f = 1
            break
        elif (row[0] == pcb2SN) and (pcb2SN == 'X'):
            snPart = row[1]
            f = 1
            break
    if f == 1:      #if recognized, unset flag
        badBoardNum = False
    else:           #if not recognized, set flag
        badBoardNum = True
               
    return [snPart, pcb2SN, badBoardNum]

#----------------------------------------------------------------------------

def genExcelSheet(sender, app_data):

    wb = load_workbook(filename = 'Data/testworkbook.xlsx')    
    ws = wb.active
    
    for idx, row in enumerate(allSingleDrives):
        copyRangeDrv('B1:J6', (1+(idx*6),2), ws)
        
        sn = ws.cell(row=(1+(idx*6)), column=9)
        sn.value = str(allSingleDrives[idx][9][0:-5])
        
        seqNum = ws.cell(row=(1+(idx*6)), column=10)
        seqNum.value = str(allSingleDrives[idx][9][-5:])

        mainNum = ws.cell(row=(2+(idx*6)), column=6)
        mainNum.value = str(allSingleDrives[idx][5])

        cpuNum = ws.cell(row=(3+(idx*6)), column=7)
        cpuNum.value = str(allSingleDrives[idx][6])

        powerNum = ws.cell(row=(4+(idx*6)), column=7)
        powerNum.value = str(allSingleDrives[idx][7])

        filterNum = ws.cell(row=(5+(idx*6)), column=7)
        filterNum.value = str(allSingleDrives[idx][8])

        fpgaNum = ws.cell(row=(3+(idx*6)), column=9)
        if (allSingleDrives[idx][1] == '325VDC') & (allSingleDrives[idx][6][0:-5] == '3202'):
            fpgaNum.value = 'FPGA: 3.2.0.5.V.B0'
        elif (allSingleDrives[idx][1] == '325VDC') & (allSingleDrives[idx][6][0:-5] == '3202A'):
            fpgaNum.value = 'FPGA: 3.3.0.5.A.B0'
        elif (allSingleDrives[idx][1] == '650VDC') & (allSingleDrives[idx][6][0:-5] == '3202A'):
            fpgaNum.value = 'FPGA: 3.3.0.5.B.B0'
        elif (allSingleDrives[idx][1] == '650VDC') & (allSingleDrives[idx][6][0:-5] == '3202'):
            fpgaNum.value = 'FPGA: 3.2.0.5.W.B0'
        else:
            print('Unrecognized voltage / CPU pcb combination; add FPGA F/W version manually.')

        dspNum = ws.cell(row=(4+(idx*6)), column=9)
        dspNum.value = 'DSP: 3.4.15.5.B.B0'
        
    t = datetime.datetime.now()
    t2 = t.strftime('%x_%X').replace(':','.')
    t2 = t2.replace('/','.')
    wb.save(dpg.get_value(outputDir)+r'\SingleDriveSN_'+t2+'.xlsx')

#----------------------------------------------------------------------------
    
def copyRangeDrv(sourceRange, destinationStartCell, sheet):
    source_cells = sheet[sourceRange]
    min_col, min_row, max_col, max_row = range_boundaries(sourceRange)
    for row_idx, row in enumerate(source_cells, start=0):
        
        for col_idx, cell in enumerate(row, start=0):
            target_cell = sheet.cell(row=destinationStartCell[0] + row_idx,
                                     column=destinationStartCell[1] + col_idx)
            
            if cell.value is not None:
                target_cell.value = cell.value
                
            if cell.has_style:
                target_cell._style = copy(cell._style)
                
    sheet.merge_cells(start_row=destinationStartCell[0], start_column=2,
                      end_row=destinationStartCell[0], end_column=5)
    sheet.merge_cells(start_row=destinationStartCell[0], start_column=6,
                      end_row=destinationStartCell[0], end_column=8)
    sheet.merge_cells(start_row=destinationStartCell[0]+1, start_column=2,
                      end_row=destinationStartCell[0]+1, end_column=5)
    sheet.merge_cells(start_row=destinationStartCell[0]+1, start_column=6,
                      end_row=destinationStartCell[0]+1, end_column=8)
    sheet.merge_cells(start_row=destinationStartCell[0]+2, start_column=2,
                      end_row=destinationStartCell[0]+2, end_column=5)
    sheet.merge_cells(start_row=destinationStartCell[0]+2, start_column=9,
                      end_row=destinationStartCell[0]+2, end_column=10)
    sheet.merge_cells(start_row=destinationStartCell[0]+3, start_column=2,
                      end_row=destinationStartCell[0]+3, end_column=5)
    sheet.merge_cells(start_row=destinationStartCell[0]+4, start_column=2,
                      end_row=destinationStartCell[0]+4, end_column=5)
    sheet.merge_cells(start_row=destinationStartCell[0]+5, start_column=2,
                      end_row=destinationStartCell[0]+5, end_column=5)

#----------------------------------------------------------------------------

def clearDrv(sender, app_data):

    allSingleDrives.pop()
    dpg.delete_item(dpg.get_value(seqNumToClear))
    
    return

#----------------------------------------------------------------------------

def clearDualDrv(sender, app_data):

    allDualDrives.pop()
    dpg.delete_item(dpg.get_value(dualSeqNumToClear))
    
    return

#-------------------------------------------------------------------------------

def serializeDual(sender, app_data):

    #########################################################################
    #Hide / show error text, depending on if bad num flags are present

    dualDrvInfo = []
    [dualDrvInfo, badDualDrvDrawingNum, badMainPcb1Num,
     badCpuPcb1Num, badMainPcb2Num, badCpuPcb2Num] = generateDualDrvSerialNumber(dpg.get_value(dualDrvNumber))

    if badDualDrvDrawingNum:
        dpg.show_item(bDDDN)
    elif dpg.is_item_shown(bDDDN):
        dpg.hide_item(bDDDN)
    
    if badMainPcb1Num:
        dpg.show_item(bM1PN)
    elif dpg.is_item_shown(bM1PN):
        dpg.hide_item(bM1PN)
   
    if badCpuPcb1Num:
        dpg.show_item(bC1PN)
    elif dpg.is_item_shown(bC1PN):
        dpg.hide_item(bC1PN)

    if badMainPcb2Num:
        dpg.show_item(bM2PN)
    elif dpg.is_item_shown(bM2PN):
        dpg.hide_item(bM2PN)
    
    if badCpuPcb2Num:
        dpg.show_item(bC2PN)
    elif dpg.is_item_shown(bC2PN):
        dpg.hide_item(bC2PN)

    #############################################################
    #If the sequential number is unique, add drive to allSingleDrives and add SN to window

    if not (badDualDrvDrawingNum or badMainPcb1Num or badCpuPcb1Num or badMainPcb2Num or badCpuPcb2Num):

        bDSQV = False        
        for row in allDualDrives:
            if (int(row[9][-4:]) == int(dpg.get_value(dualSeqNum))):
                bDSQV = True
                break

        if not bDSQV:
            allDualDrives.append(dualDrvInfo)
            dpg.add_text(
                default_value=dualDrvInfo[9],
                parent=dualSnWindow,
                tag=dpg.get_value(dualSeqNum) 
            )
            dpg.hide_item(bDSQN)
        else:
            dpg.show_item(bDSQN)
            
    return

#-------------------------------------------------------------------------------

def generateDualDrvSerialNumber(drvNum):

    [dualDrvInfo, badDualDrvDrawingNum] = getDualDriveInfo(drvNum)

    boardNums = [dpg.get_value(dual1MainPCB), dpg.get_value(dual1CpuPCB),
                 dpg.get_value(dual2MainPCB), dpg.get_value(dual2CpuPCB)]
    dualDrvInfo.extend(boardNums)

    if dualDrvInfo[2] == "inverted":                            #first two places
        sn = "DH"
    elif dualDrvInfo[2] == "upright":
        sn = "DF"

    sn = sn + "04"      #3-4 places

    [snPartA, dualDrvInfo[5], badMainPcb1Num] = serializeFromCSV(r'Data\snDualDriveMainBoard.csv', 'Main 1', dualDrvInfo[5]) #5-6
    sn = sn + snPartA

    [snPartB, dualDrvInfo[6], badCpuPcb1Num] = serializeFromCSV(r'Data\snDualDriveCPUBoard.csv', 'CPU 1', dualDrvInfo[6]) #7
    sn = sn + snPartB

    [snPartC, dualDrvInfo[7], badMainPcb2Num] = serializeFromCSV(r'Data\snDualDriveMainBoard.csv', 'Main 2', dualDrvInfo[7]) #8-9
    sn = sn + snPartC
        
    [snPartD, dualDrvInfo[8], badCpuPcb2Num] = serializeFromCSV(r'Data\snDualDriveCpuBoard.csv','CPU 2', dualDrvInfo[8]) #10
    sn = sn + snPartD

    sn = sn +'-'+ dpg.get_value(dualSeqNum)  
    dualDrvInfo.append(str(sn))
        
    return [dualDrvInfo, badDualDrvDrawingNum, badMainPcb1Num, badCpuPcb1Num, badMainPcb2Num, badCpuPcb2Num]

#-------------------------------------------------------------------------------

def getDualDriveInfo(drvNum):

##############################################################
#load the reference CSV file

    with open(r'Data/dualDriveNumbers.csv', 'r', newline='', encoding='utf-8-sig') as csvfile:
        csv_reader = csv.reader(csvfile, delimiter=',', quotechar='|')
        driveNumbers = list(csv_reader)
        csvfile.close()
        
##############################################################
#Loop through all entries in the file

    drvInfo = ['43687REVC','650VDC','inverted','25A','single']
    f = 0     

    for row in driveNumbers:
        if row[0] == drvNum:
            drvInfo = row
            f = 1
            break
    
    if f == 1:      #if # is recognized, unset flag
        badDrvDrawingNum = False
    else:           #if # is NOT recognized, setting flag that brings up error text
        badDrvDrawingNum = True

    return [drvInfo, badDrvDrawingNum]

#----------------------------------------------------------------------------

def genExcelSheetDual(sender, app_data):

    wb = load_workbook(filename = 'Data/testworkbookDual.xlsx')    
    ws = wb.active
    
    for idx, row in enumerate(allDualDrives):
        copyRangeDualDrv('B1:J6', (1+(idx*6),2), ws)
        
        sn = ws.cell(row=(1+(idx*6)), column=9)
        sn.value = str(allDualDrives[idx][9][0:-5])
        
        seqNum = ws.cell(row=(1+(idx*6)), column=10)
        seqNum.value = str(allDualDrives[idx][9][-5:])

        main1Num = ws.cell(row=(3+(idx*6)), column=7)
        main1Num.value = str(allDualDrives[idx][5])

        cpu1Num = ws.cell(row=(2+(idx*6)), column=7)
        cpu1Num.value = str(allDualDrives[idx][6])

        main2Num = ws.cell(row=(5+(idx*6)), column=7)
        main2Num.value = str(allDualDrives[idx][7])

        cpu2Num = ws.cell(row=(4+(idx*6)), column=7)
        cpu2Num.value = str(allDualDrives[idx][8])

        fpgaNum = ws.cell(row=(2+(idx*6)), column=9)
        if (allDualDrives[idx][1] == '325VDC') & (allDualDrives[idx][6][0:-5] == '3202'):
            fpgaNum.value = 'FPGA: 3.2.0.5.V.B0'
        elif (allDualDrives[idx][1] == '325VDC') & (allDualDrives[idx][6][0:-5] == '3202A'):
            fpgaNum.value = 'FPGA: 3.3.0.5.A.B0'
        elif (allDualDrives[idx][1] == '650VDC') & (allDualDrives[idx][6][0:-5] == '3202A'):
            fpgaNum.value = 'FPGA: 3.3.0.5.B.B0'
        elif (allDualDrives[idx][1] == '650VDC') & (allDualDrives[idx][6][0:-5] == '3202'):
            fpgaNum.value = 'FPGA: 3.2.0.5.W.B0'
        else:
            print('Unrecognized voltage / CPU pcb combination; add FPGA F/W version manually.')

        dspNum = ws.cell(row=(3+(idx*6)), column=9)
        dspNum.value = 'DSP: 3.4.15.5.B.B0'
        
    t = datetime.datetime.now()
    t2 = t.strftime('%x_%X').replace(':','.')
    t2 = t2.replace('/','.')
    wb.save(dpg.get_value(dualOutputDir)+r'\DualDriveSN_'+t2+'.xlsx')

#----------------------------------------------------------------------------
    
def copyRangeDualDrv(sourceRange, destinationStartCell, sheet):
    source_cells = sheet[sourceRange]
    min_col, min_row, max_col, max_row = range_boundaries(sourceRange)
    for row_idx, row in enumerate(source_cells, start=0):
        
        for col_idx, cell in enumerate(row, start=0):
            target_cell = sheet.cell(row=destinationStartCell[0] + row_idx,
                                     column=destinationStartCell[1] + col_idx)
            
            if cell.value is not None:
                target_cell.value = cell.value
                
            if cell.has_style:
                target_cell._style = copy(cell._style)
                
    sheet.merge_cells(start_row=destinationStartCell[0], start_column=2,
                      end_row=destinationStartCell[0], end_column=5)
    sheet.merge_cells(start_row=destinationStartCell[0], start_column=6,
                      end_row=destinationStartCell[0], end_column=8)
    sheet.merge_cells(start_row=destinationStartCell[0]+1, start_column=2,
                      end_row=destinationStartCell[0]+1, end_column=5)
    sheet.merge_cells(start_row=destinationStartCell[0]+2, start_column=2,
                      end_row=destinationStartCell[0]+2, end_column=5)
    sheet.merge_cells(start_row=destinationStartCell[0]+1, start_column=9,
                      end_row=destinationStartCell[0]+1, end_column=10)
    sheet.merge_cells(start_row=destinationStartCell[0]+3, start_column=2,
                      end_row=destinationStartCell[0]+3, end_column=5)
    sheet.merge_cells(start_row=destinationStartCell[0]+4, start_column=2,
                      end_row=destinationStartCell[0]+4, end_column=5)
    sheet.merge_cells(start_row=destinationStartCell[0]+5, start_column=2,
                      end_row=destinationStartCell[0]+5, end_column=5)

#-------------------------------------------------------------------------

def getPsInfo(psNum):

##############################################################
#load the reference CSV file

    with open(r'Data/psNumbers.csv', 'r', newline='', encoding='utf-8-sig') as csvfile:
        csv_reader = csv.reader(csvfile, delimiter=',', quotechar='|')
        psNumbers = list(csv_reader)
        csvfile.close()
 
##############################################################
#Loop through all entries in the file
    
    psInfo = ['43695REVF','inverted','380-480 VAC','540-680 VDC']
    f = 0
         
    for row in psNumbers:
        if row[0] == psNum:
            psInfo = row
            f = 1
            break
            
    if f == 1:      #if # is recognized, unset flag
        badPsDrawingNum = False
    else:           #if # is NOT recognized, setting flag that brings up error text
        badPsDrawingNum = True
            
    return [psInfo, badPsDrawingNum]

#-------------------------------------------------------------------------

def generatePsSerialNumber(psNum):

##############################################################
#get the drawing and pcb numbers

    [psInfo, badPsDrawingNum] = getPsInfo(dpg.get_value(psNumber))
    psInfo.append(str(dpg.get_value(psMainPCB)))

    sn = "PS"                   #1 - 2

    if psInfo[1] == "inverted":     #3
        sn = sn + "C"
    elif psInfo[1] == "upright":
        sn = sn + "A"

    sn = sn + "04"              #4

    [snPartA, psInfo[4], badPsMainPcbNum] = serializeFromCSV(r'Data\snPSMainBoard.csv', 'Main', psInfo[4]) #5-6
    sn = sn + snPartA

    sn = sn +'-'+ dpg.get_value(psSeqNum)
    psInfo.append(str(sn))

    return [psInfo, badPsDrawingNum, badPsMainPcbNum]

#-------------------------------------------------------------------------------

def serializePS(sender, app_data):

    #########################################################################
    #Hide / show error text, depending on if bad num flags are present

    psInfo = []
    [psInfo, badPsDrawingNum, badPsMainPcbNum] = generatePsSerialNumber(dpg.get_value(psNumber))

    if badPsDrawingNum:
        dpg.show_item(bPSDN)
    elif dpg.is_item_shown(bPSDN):
        dpg.hide_item(bPSDN)
    
    if badPsMainPcbNum:
        dpg.show_item(bPSMN)
    elif dpg.is_item_shown(bPSMN):
        dpg.hide_item(bPSMN)
   
    #############################################################
    #If the sequential number is unique, add drive to allSingleDrives and add SN to window

    if not (badPsDrawingNum or badPsMainPcbNum):

        bPSQV = False        
        for row in allPowerSupplies:
            if (int(row[5][-4:]) == int(dpg.get_value(psSeqNum))):
                bPSQV = True
                break

        if not bPSQV:
            allPowerSupplies.append(psInfo)
            dpg.add_text(
                default_value=psInfo[5],
                parent=psSnWindow,
                tag=dpg.get_value(psSeqNum) 
            )
            dpg.hide_item(bPSQN)
        else:
            dpg.show_item(bPSQN)
            
    return

#----------------------------------------------------------------------------

def clearPs(sender, app_data):

    allPowerSupplies.pop()
    dpg.delete_item(dpg.get_value(psSeqNumToClear))
    
    return

#----------------------------------------------------------------------------

def genExcelSheetPS(sender, app_data):

    wb = load_workbook(filename = 'Data/testworkbookPS.xlsx')    
    ws = wb.active
    
    for idx, row in enumerate(allPowerSupplies):
        copyRangePS('B1:J3', (1+(idx*3),2), ws)
        
        sn = ws.cell(row=(1+(idx*3)), column=6)
        sn.value = str(allPowerSupplies[idx][5][0:-5])
        
        seqNum = ws.cell(row=(1+(idx*3)), column=10)
        seqNum.value = str(allPowerSupplies[idx][5][-5:])

        mainNum = ws.cell(row=(2+(idx*3)), column=6)
        mainNum.value = str(allPowerSupplies[idx][4])
        
    t = datetime.datetime.now()
    t2 = t.strftime('%x_%X').replace(':','.')
    t2 = t2.replace('/','.')
    wb.save(dpg.get_value(psOutputDir)+r'\PowerSupplySN_'+t2+'.xlsx')

#----------------------------------------------------------------------------
    
def copyRangePS(sourceRange, destinationStartCell, sheet):
    source_cells = sheet[sourceRange]
    min_col, min_row, max_col, max_row = range_boundaries(sourceRange)
    for row_idx, row in enumerate(source_cells, start=0):
        
        for col_idx, cell in enumerate(row, start=0):
            target_cell = sheet.cell(row=destinationStartCell[0] + row_idx,
                                     column=destinationStartCell[1] + col_idx)
            
            if cell.value is not None:
                target_cell.value = cell.value
                
            if cell.has_style:
                target_cell._style = copy(cell._style)
                
    sheet.merge_cells(start_row=destinationStartCell[0], start_column=2,
                      end_row=destinationStartCell[0], end_column=5)
    sheet.merge_cells(start_row=destinationStartCell[0], start_column=6,
                      end_row=destinationStartCell[0], end_column=9)
    sheet.merge_cells(start_row=destinationStartCell[0]+1, start_column=2,
                      end_row=destinationStartCell[0]+1, end_column=5)
    sheet.merge_cells(start_row=destinationStartCell[0]+2, start_column=2,
                      end_row=destinationStartCell[0]+2, end_column=3)
    sheet.merge_cells(start_row=destinationStartCell[0]+1, start_column=6,
                      end_row=destinationStartCell[0]+1, end_column=10)

#----------------------------------------------------------------------------

with dpg.window(width=750, height=600, pos=[0,0], no_title_bar=True):
    
    tabBar = dpg.add_tab_bar(
    )
    drvTab = dpg.add_tab(
        label='Drives',
        parent=tabBar
    )
    dualDrvTab = dpg.add_tab(
        label='Dual Drives',
        parent=tabBar
    )
    psTab = dpg.add_tab(
        label='Power Supplies',
        parent=tabBar
    )

    drvNumber = dpg.add_input_text(
        label="Drive Drawing Number?",
        hint='#####, or #####REVX',
        width=300,
        uppercase=True,
        no_spaces=True,
        parent=drvTab
    )
    bDDN = dpg.add_text(
        default_value='Incorrect or Unknown Drive Drawing Number.',
        color=[255,0,0],
        parent=drvTab,
        show=False
    )

    
    mainPCB = dpg.add_input_text(
        label="Main PCB Number?",
        hint='####-####, or ####X-####',
        width=300,
        uppercase=True,
        no_spaces=True,
        parent=drvTab
    )
    bMPN = dpg.add_text(
        default_value='Incorrect or Unknown Main PCB Number.',
        color=[255,0,0],
        parent=drvTab,
        show=False
    )

    
    cpuPCB = dpg.add_input_text(
        label="CPU PCB Number?",
        hint='####-####, or ####X-####',
        width=300,
        uppercase=True,
        no_spaces=True,
        parent=drvTab
    )
    bCPN = dpg.add_text(
        default_value='Incorrect or Unknown CPU PCB Number.',
        color=[255,0,0],
        parent=drvTab,
        show=False
    )

    
    powerPCB = dpg.add_input_text(
        label="Power PCB Number?",
        hint='####-####, or ####X-####, OR Enter "X"',
        width=300,
        uppercase=True,
        no_spaces=True,
        parent=drvTab
    )
    bPPN = dpg.add_text(
        default_value='Incorrect or Unknown Power PCB Number.',
        color=[255,0,0],
        parent=drvTab,
        show=False
    )

    
    filterPCB = dpg.add_input_text(
        label="Filter PCB Number?",
        hint='####-####, or ####X-####, OR Enter "X"',
        width=300,
        uppercase=True,
        no_spaces=True,
        parent=drvTab
    )
    bFPN = dpg.add_text(
        default_value='Incorrect or Unknown Filter PCB Number.',
        color=[255,0,0],
        parent=drvTab,
        show=False
    )

    
    seqNum = dpg.add_input_text(
        label="Sequential Number?",
        hint='####',
        width=300,
        decimal=True,
        no_spaces=True,
        parent=drvTab,
        tag='seqNum'
    )
    bSQN = dpg.add_text(
        default_value='Enter a Unique Sequential Number',
        color=[255,0,0],
        parent=drvTab,
        show=False
    )
    drvSerializeButton = dpg.add_button(
        label="Serialize",
        width=100,
        callback=serialize,
        parent=drvTab
    )

    
    snWindow = dpg.add_child_window(
        width=200,
        height=400,
        pos=[500,30],
        parent=drvTab
    )
    seqNumToClear = dpg.add_input_text(
        label="Sequential Number to Clear?",
        hint='####',
        width=40,
        decimal=True,
        no_spaces=True,
        parent=drvTab,
        pos=[480,435]
    )
    drvClearButton = dpg.add_button(
        label="Clear",
        width=100,
        callback=clearDrv,
        parent=drvTab,
        pos=[480,460]
    )


    outputDir = dpg.add_input_text(
        label="Output Directory?",
        hint=r'\...\...',
        width=300,
        parent=drvTab
    )
    drvExportButton = dpg.add_button(
        label="Export",
        width=100,
        callback=genExcelSheet,
        parent=drvTab
    )

    
    dualDrvNumber = dpg.add_input_text(
        label="Dual Drive Drawing Number?",
        hint='#####, or #####REVX',
        width=300,
        uppercase=True,
        no_spaces=True,
        parent=dualDrvTab
    )
    bDDDN = dpg.add_text(
        default_value='Incorrect or Unknown Dual Drive Drawing Number.',
        color=[255,0,0],
        parent=dualDrvTab,
        show=False
    )

    
    dual1MainPCB = dpg.add_input_text(
        label="Main PCB 1 Number?",
        hint='####-####, or ####X-####',
        width=300,
        uppercase=True,
        no_spaces=True,
        parent=dualDrvTab
    )
    bM1PN = dpg.add_text(
        default_value='Incorrect or Unknown Main PCB 1 Number.',
        color=[255,0,0],
        parent=dualDrvTab,
        show=False
    )

    
    dual1CpuPCB = dpg.add_input_text(
        label="CPU PCB 1 Number?",
        hint='####-####, or ####X-####',
        width=300,
        uppercase=True,
        no_spaces=True,
        parent=dualDrvTab
    )
    bC1PN = dpg.add_text(
        default_value='Incorrect or Unknown CPU PCB 1 Number.',
        color=[255,0,0],
        parent=dualDrvTab,
        show=False
    )


    dual2MainPCB = dpg.add_input_text(
        label="Main PCB 2 Number?",
        hint='####-####, or ####X-####',
        width=300,
        uppercase=True,
        no_spaces=True,
        parent=dualDrvTab
    )
    bM2PN = dpg.add_text(
        default_value='Incorrect or Unknown Main PCB 2 Number.',
        color=[255,0,0],
        parent=dualDrvTab,
        show=False
    )

    
    dual2CpuPCB = dpg.add_input_text(
        label="CPU PCB 2 Number?",
        hint='####-####, or ####X-####',
        width=300,
        uppercase=True,
        no_spaces=True,
        parent=dualDrvTab
    )
    bC2PN = dpg.add_text(
        default_value='Incorrect or Unknown CPU PCB 2 Number.',
        color=[255,0,0],
        parent=dualDrvTab,
        show=False
    )

    dualSeqNum = dpg.add_input_text(
        label="Sequential Number?",
        hint='####',
        width=300,
        decimal=True,
        no_spaces=True,
        parent=dualDrvTab,
        tag='dualSeqNum'
    )
    bDSQN = dpg.add_text(
        default_value='Enter a Unique Sequential Number',
        color=[255,0,0],
        parent=dualDrvTab,
        show=False
    )
    dualDrvSerializeButton = dpg.add_button(
        label="Serialize",
        width=100,
        callback=serializeDual,
        parent=dualDrvTab
    )


    dualSnWindow = dpg.add_child_window(
        width=200,
        height=400,
        pos=[500,30],
        parent=dualDrvTab
    )
    dualSeqNumToClear = dpg.add_input_text(
        label="Sequential Number to Clear?",
        hint='####',
        width=40,
        decimal=True,
        no_spaces=True,
        parent=dualDrvTab,
        pos=[480,435]
    )
    dualDrvClearButton = dpg.add_button(
        label="Clear",
        width=100,
        callback=clearDualDrv,
        parent=dualDrvTab,
        pos=[480,460]
    )

    dualOutputDir = dpg.add_input_text(
        label="Output Directory?",
        hint=r'\...\...',
        width=300,
        parent=dualDrvTab
    )
    dualDrvExportButton = dpg.add_button(
        label="Export",
        width=100,
        callback=genExcelSheetDual,
        parent=dualDrvTab
    )


    psNumber = dpg.add_input_text(
        label="Power Supply Drawing Number?",
        hint='#####, or #####REVX',
        width=300,
        uppercase=True,
        no_spaces=True,
        parent=psTab
    )
    bPSDN = dpg.add_text(
        default_value='Incorrect or Unknown Power Supply Drawing Number.',
        color=[255,0,0],
        parent=psTab,
        show=False
    )


    psMainPCB = dpg.add_input_text(
        label="PCB Number?",
        hint='###-####, or ###X-####',
        width=300,
        uppercase=True,
        no_spaces=True,
        parent=psTab
    )
    bPSMN = dpg.add_text(
        default_value='Incorrect or Unknown Main PCB Number.',
        color=[255,0,0],
        parent=psTab,
        show=False
    )


    psSeqNum = dpg.add_input_text(
        label="Sequential Number?",
        hint='####',
        width=300,
        decimal=True,
        no_spaces=True,
        parent=psTab,
        tag='psSeqNum'
    )
    bPSQN = dpg.add_text(
        default_value='Enter a Unique Sequential Number',
        color=[255,0,0],
        parent=psTab,
        show=False
    )
    psSerializeButton = dpg.add_button(
        label="Serialize",
        width=100,
        callback=serializePS,
        parent=psTab
    )

    psSnWindow = dpg.add_child_window(
        width=200,
        height=400,
        pos=[500,30],
        parent=psTab
    )
    psSeqNumToClear = dpg.add_input_text(
        label="Sequential Number to Clear?",
        hint='####',
        width=40,
        decimal=True,
        no_spaces=True,
        parent=psTab,
        pos=[480,435]
    )
    psClearButton = dpg.add_button(
        label="Clear",
        width=100,
        callback=clearPs,
        parent=psTab,
        pos=[480,460]
    )

    psOutputDir = dpg.add_input_text(
        label="Output Directory?",
        hint=r'\...\...',
        width=300,
        parent=psTab
    )
    psExportButton = dpg.add_button(
        label="Export",
        width=100,
        callback=genExcelSheetPS,
        parent=psTab
    )
dpg.create_viewport(title='Serialize ACS Drives', width=750, height=600)
dpg.setup_dearpygui()
dpg.show_viewport()
dpg.start_dearpygui()
dpg.destroy_context()

#-------------------------------------------------------------------------------

