import csv
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from copy import copy
import datetime

def serialize():

##############################################################
#get amount of drives

    print("Enter the number of drives to serialize")
    driveCount = int(input())

##############################################################
#get amount of power supplies

    print("Enter the number of power supplies to serialize")
    psCount = int(input())

##############################################################
#loop through # of drives, add list with drive information to a list

    i = 0
    drvSerials = []
    
    while i < driveCount:
        
#[drawingNum, Voltage, U or I, Current, Single or Dual, Main pcb, CPU pcb,
# Power pcb, Filter pcb, Serial Num]

        drvSerials.append(generateDriveSerialNumber())           
        i = i + 1

    genExcelSheet(drvSerials, '', 4422)

    i = 0
    psSerials = []
    
    while i < psCount:
        
        psSerials.append(generatePSSerialNumber())
        i += 1
        
    print(psSerials)
#-------------------------------------------------------------------------

def getPSDrawingNumber():

##############################################################
#get the PS drawing number

    print("Power Supply Drawing # (Format: ##### or #####REVX):")
    psNum = input()

##############################################################
#load the reference CSV file

    with open(r'Data/psNumbers.csv', 'r', newline='', encoding='utf-8-sig') as csvfile:
        csv_reader = csv.reader(csvfile, delimiter=',', quotechar='|')
        psNumbers = list(csv_reader)
        csvfile.close()
        
##############################################################
#Loop through all entries in the file, if number isn't recognized, 
#loop through asking for number to be re-entered or to add a new entry to the file

        f = 0

        while f == 0:
         
            for row in psNumbers:
                if row[0] == psNum:
                    psInfo = row
                    f = 1
                    break
            
            if f == 1:      #break loop if # is recognized
                break
            
            print("Unknown or incorrect Power Supply Drawing number. Re-enter number, "
                  "or enter an 'x' to add a new Power Supply Drawing number entry.")
            psNum = input()
            
            if psNum == "x":        #if selected, write new line to CSV
                with open(r'Data/psNumbers.csv', 'a', newline='', encoding='utf-8-sig') as csvfile:
                    psNumbersWriter = csv.writer(csvfile, delimiter=',', quotechar='|')
                    psInfo = [input("PS Number?"),input("'upright' or 'inverted'?"),
                          "380-480 VAC", "540-680 VDC"]                     
                    psNumbersWriter.writerow(psInfo)
                    csvfile.close()
                break

    return psInfo

#-------------------------------------------------------------------------

def getPCBNumbersPS():

##############################################################
#get 1st main board number
    
    print("Power Supply PCB Number:")
    mainBoard = input()

    return mainBoard

#-------------------------------------------------------------------------
            

def getPCBNumbersSingleNode():

##############################################################
#get main board number
    
    print("Main PCB Number:")
    mainBoard = input()

##############################################################
#get CPU board number
    
    print("CPU PCB Number:")
    cpuBoard = input()

##############################################################
#get Power board number, if present
    
    print("Power PCB Number (enter 'X' if not present:")
    powerBoard = input()

##############################################################
#get Filter board number, if present
    
    print("Filter PCB Number (enter 'X' if not present:")
    filterBoard = input()

    boardNums = [mainBoard, cpuBoard, powerBoard, filterBoard]
    
    return boardNums

#-------------------------------------------------------------------------

def getDriveDrawingNumber():

##############################################################
#get the drive number
    
    print("Drive Drawing # (Format: ##### or #####REVX):")
    drvNum = input()

##############################################################
#load the reference CSV file

    with open(r'Data/driveNumbers.csv', 'r', newline='', encoding='utf-8-sig') as csvfile:
        csv_reader = csv.reader(csvfile, delimiter=',', quotechar='|')
        driveNumbers = list(csv_reader)
        csvfile.close()
        
##############################################################
#Loop through all entries in the file, if number isn't recognized, 
#loop through asking for number to be re-entered or to add a new entry to the file

    f = 0

    while f == 0:
         
        for row in driveNumbers:
            if row[0] == drvNum:
                drvInfo = row
                f = 1
                break
            
        if f == 1:      #break loop if # is recognized
            break
            
        print("Unknown or incorrect Drive Drawing number. Re-enter number, "
                  "or enter an 'n' to add a new Drive Drawing number entry.")
        drvNum = input()
            
        if drvNum == "n":        #if selected, write new line to CSV
            with open(r'Data/driveNumbers.csv', 'a', newline='', encoding='utf-8-sig') as csvfile:
                drvNumbersWriter = csv.writer(csvfile, delimiter=',', quotechar='|')
                drvInfo = [input("Drive Drawing Number?: "), input("Voltage? (650VDC or 325VDC): "),
                            input("'upright' or 'inverted'?: "), input("Current rating?: "),
                            input("'single' or 'dual' node?: ")]                     
                drvNumbersWriter.writerow(drvInfo)
                csvfile.close()
            break

    return drvInfo

#-------------------------------------------------------------------------

def getPCBNumbersDualNode():

##############################################################
#get 1st main board number
    
    print("Main PCB Number 1:")
    mainBoard1 = input()

##############################################################
#get 1st CPU board number
    
    print("CPU PCB Number 1:")
    cpuBoard1 = input()

##############################################################
#get 2nd main board number
    
    print("Main PCB Number 2:")
    mainBoard2 = input()

##############################################################
#get 2nd CPU board number
    
    print("CPU PCB Number 2:")
    cpuBoard2 = input()
    
    boardNums = [mainBoard1, cpuBoard1, mainBoard2, cpuBoard2]
    
    return boardNums

#-------------------------------------------------------------------------

def generatePSSerialNumber():

##############################################################
#get the drawing and pcb numbers

    psInfo = getPSDrawingNumber()
    psInfo.extend(getPCBNumbersPS())
    
##############################################################
#Generate serial number in probably a really stupid way but for power supplies

    sn = "PS"                   #1 - 2

    if psInfo[1] == "inverted":     #3
        sn = sn + "C"
    elif psInfo[1] == "upright":
        sn = sn + "A"

    sn = sn + "04"              #4

    [snPartA, psInfo[4]] = serializeFromCSV(r'Data\snPSMainBoard.csv', 'Main', psInfo[4]) #5-6
    sn = sn + snPartA

    psInfo.append(str(sn))

    return psInfo

    
#----------------------------------------------------------------------------

def generateDriveSerialNumber():

##############################################################
#get the drive number

    drvInfo = getDriveDrawingNumber()
    
##############################################################
#get the pcb numbers, depending on whether drive is dual or single node

    if drvInfo[4] == "single":
        boardNums = getPCBNumbersSingleNode()
        drvInfo.extend(boardNums)
    
##############################################################
#Generate serial number in probably a really stupid way but I'm too stupid
#to do it another way        

        if drvInfo[2] == "inverted":                            #first two places
            sn = "DG"
        elif drvInfo[2] == "upright":
            sn = "DE"

        if drvInfo[1] == "650VDC" and drvInfo[3] == "13A":      #3-4 places
            sn = sn + "10"                                      
        elif drvInfo[1] == "650VDC" and drvInfo[3] == "25A":
            sn = sn + "06"
        elif drvInfo[1] == "650VDC" and drvInfo[3] == "50A":
            sn = sn + "13"
        elif drvInfo[1] == "325VDC" and drvInfo[3] == "25A":
            sn = sn + "14"
        elif drvInfo[1] == "325VDC" and drvInfo[3] == "50A":
            sn = sn + "12"
        elif drvInfo[1] == "325VDC" and drvInfo[3] == "100A":
            sn = sn + "11"

        [snPartA, drvInfo[5]] = serializeFromCSV(r'Data\snDriveMainBoard.csv', 'Main', drvInfo[5]) #5-6
        sn = sn + snPartA

        [snPartB, drvInfo[6]] = serializeFromCSV(r'Data\snDriveCPUBoard.csv', 'CPU', drvInfo[6]) #7
        sn = sn + snPartB

        [snPartC, drvInfo[7]] = serializeFromCSV(r'Data\snDrivePowerBoard.csv', 'Power', drvInfo[7]) #8-9
        sn = sn + snPartC
        
        [snPartD, drvInfo[8]] = serializeFromCSV(r'Data\snDriveFilterBoard.csv','Filter', drvInfo[8]) #10
        sn = sn + snPartD
  
        drvInfo.append(str(sn))
        
##############################################################
#get the pcb numbers for dual node

    elif drvInfo[4] == "dual":
        boardNums = getPCBNumbersDualNode()
        drvInfo.extend(boardNums)

##############################################################
#Generate serial number in probably a really stupid way but for dual node drives

        if drvInfo[2] == "inverted":                            #first two places
            sn = "DH"
        elif drvInfo[2] == "upright":
            sn = "DF"

        sn = sn + "04"                                          #3-4

        [snPartA, drvInfo[6]] = serializeFromCSV(r'Data\snDriveCPUBoard.csv', '1st CPU', drvInfo[6]) #5
        sn = sn + snPartA

        [snPartB, drvInfo[5]] = serializeFromCSV(r'Data\snDualDriveMainBoard.csv', '1st Main', drvInfo[5]) #6-7
        sn = sn + snPartB

        [snPartC, drvInfo[8]] = serializeFromCSV(r'Data\snDriveCPUBoard.csv', '2nd CPU', drvInfo[8]) #8
        sn = sn + snPartC

        [snPartD, drvInfo[7]] = serializeFromCSV(r'Data\snDualDriveMainBoard.csv', '2nd Main', drvInfo[7]) #9-10
        sn = sn + snPartD 

        drvInfo.append(str(sn))
        
    else:
        print("*kazoo noises*")

    return drvInfo

#----------------------------------------------------------------------------
        
def serializeFromCSV(filename, pcbType, pcb2SN):       #string filename, pcb to serialize (####X-####), 
                                                      #type of PCB for print statements
##############################################################
#load the reference CSV file

        with open(filename, 'r', newline='', encoding='utf-8-sig') as csvfile:
            csv_reader = csv.reader(csvfile, delimiter=',', quotechar='|')
            boards = list(csv_reader)
        csvfile.close()
            
##############################################################
#Loop through all entries in the file, if number isn't recognized, 
#loop through asking for number to be re-entered or to add a new entry to the file

        f = 0
        while f == 0:
             
                for row in boards:
                    if row[0] == pcb2SN[0:-5]:
                        snPart = row[1]
                        f = 1
                        break
                    elif row[0] == pcb2SN:
                        snPart = row[1]
                        f = 1
                        break
                if f == 1:      #break loop if # is recognized
                    break
                
                print("Unknown or incorrect "+ pcbType +" PCB number. Re-enter number, (Format: ####X-####)"
                      "or enter an 'n' to add a new "+ pcbType +" PCB number entry.")
                bnum = input()
                
                if bnum == "n":        #if selected, write new line to CSV
                    with open(filename, 'a', newline='', encoding='utf-8-sig') as csvfile:
                        pcbWriter = csv.writer(csvfile, delimiter=',', quotechar='|')
                        pcbInfo = [input(pcbType +" PCB Number? (Format: ####X-####): "),
                                   input("Corresponding serial number portion: ")]
                        
                        bnum = pcbInfo[0]                   #save non stripped version to variable that 
                        pcbInfo[0] = pcbInfo[0][0:-5]       #write back the stripped version to the write to csv file
                        pcbWriter.writerow(pcbInfo)
                        csvfile.close()

                    pcb2SN = bnum
                    snPart = pcbInfo[1]
                    break

                pcb2SN = bnum

        return [snPart, pcb2SN]

#----------------------------------------------------------------------------

def genExcelSheet(drvInfo, outputDirectory, startingNum):

    wb = load_workbook(filename = 'Data/testworkbook.xlsx')    
    ws = wb.active

    for idx, row in enumerate(drvInfo):
        copyRange('B1:J6', (1+(idx*6),2), ws)
        
        sn = ws.cell(row=(1+(idx*6)), column=9)
        sn.value = str(drvInfo[idx][9])
        
        seqNum = ws.cell(row=(1+(idx*6)), column=10)
        seqNum.value = '- '+str(startingNum+idx)

        mainNum = ws.cell(row=(2+(idx*6)), column=6)
        mainNum.value = str(drvInfo[idx][5])

        cpuNum = ws.cell(row=(3+(idx*6)), column=7)
        cpuNum.value = str(drvInfo[idx][6])

        powerNum = ws.cell(row=(4+(idx*6)), column=7)
        powerNum.value = str(drvInfo[idx][7])

        filterNum = ws.cell(row=(5+(idx*6)), column=7)
        filterNum.value = str(drvInfo[idx][8])

        fpgaNum = ws.cell(row=(3+(idx*6)), column=9)
        if (drvInfo[idx][1] == '325VDC') & (drvInfo[idx][6][0:-5] == '3202'):
            fpgaNum.value = 'FPGA: 3.2.0.5.V.B0'
        elif (drvInfo[idx][1] == '325VDC') & (drvInfo[idx][6][0:-5] == '3202A'):
            fpgaNum.value = 'FPGA: 3.3.0.5.A.B0'
        elif (drvInfo[idx][1] == '650VDC') & (drvInfo[idx][6][0:-5] == '3202A'):
            fpgaNum.value = 'FPGA: 3.3.0.5.B.B0'
        elif (drvInfo[idx][1] == '650VDC') & (drvInfo[idx][6][0:-5] == '3202'):
            fpgaNum.value = 'FPGA: 3.2.0.5.W.B0'
        else:
            print('Unrecognized voltage / CPU pcb combination; add FPGA F/W version manually.')

        dspNum = ws.cell(row=(4+(idx*6)), column=9)
        dspNum.value = 'DSP: 3.4.15.5.B.B0'
        
    t = datetime.datetime.now()
    t2 = t.strftime('%x_%X').replace(':','.')
    t2 = t2.replace('/','.')
    wb.save(outputDirectory+'SN_'+t2+'.xlsx')

#----------------------------------------------------------------------------
    
def copyRange(sourceRange, destinationStartCell, sheet):
    source_cells = sheet[sourceRange]
    min_col, min_row, max_col, max_row = range_boundaries(sourceRange)
    for row_idx, row in enumerate(source_cells, start=0):
        for col_idx, cell in enumerate(row, start=0):
            target_cell = sheet.cell(row=destinationStartCell[0] + row_idx, column=destinationStartCell[1] + col_idx)
            if cell.value is not None:
                target_cell.value = cell.value
            if cell.has_style:
                target_cell._style = copy(cell._style)
    sheet.merge_cells(start_row=destinationStartCell[0], start_column=2, end_row=destinationStartCell[0], end_column=5)
    sheet.merge_cells(start_row=destinationStartCell[0], start_column=6, end_row=destinationStartCell[0], end_column=8)
    sheet.merge_cells(start_row=destinationStartCell[0]+1, start_column=2, end_row=destinationStartCell[0]+1, end_column=5)
    sheet.merge_cells(start_row=destinationStartCell[0]+1, start_column=6, end_row=destinationStartCell[0]+1, end_column=8)
    sheet.merge_cells(start_row=destinationStartCell[0]+2, start_column=2, end_row=destinationStartCell[0]+2, end_column=5)
    sheet.merge_cells(start_row=destinationStartCell[0]+2, start_column=9, end_row=destinationStartCell[0]+2, end_column=10)
    sheet.merge_cells(start_row=destinationStartCell[0]+3, start_column=2, end_row=destinationStartCell[0]+3, end_column=5)
    sheet.merge_cells(start_row=destinationStartCell[0]+4, start_column=2, end_row=destinationStartCell[0]+4, end_column=5)
    sheet.merge_cells(start_row=destinationStartCell[0]+5, start_column=2, end_row=destinationStartCell[0]+5, end_column=5)

